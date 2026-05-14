from django.contrib import admin
from django.utils.html import format_html
from django.urls import reverse
from django.utils.translation import gettext_lazy as _
from django.db.models import Sum, Count
import csv
import io
from django.http import HttpResponse
from django.utils import timezone
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from .models import (
    Plan, PlanFeature, Price, Subscription, SubscriptionHistory,
    Invoice, InvoiceLineItem, Payment, PaymentMethod,
    WebhookEvent, QuotaLimit, QuotaUsage
)


class BaseBillingAdmin(admin.ModelAdmin):
    list_per_page = 50
    date_hierarchy = 'created_at'
    readonly_fields = ('id', 'created_at', 'updated_at', 'created_by', 'updated_by')
    def save_model(self, request, obj, form, change):
        if not change:
            obj.created_by = request.user
        obj.updated_by = request.user
        super().save_model(request, obj, form, change)

@admin.register(Plan)
class PlanAdmin(BaseBillingAdmin):
    list_display = ('name', 'plan_type', 'price_monthly', 'price_yearly', 'currency', 'trial_days', 'is_active', 'display_order')
    list_filter = ('plan_type', 'is_active', 'currency')
    search_fields = ('name', 'slug', 'description')
    list_editable = ('display_order', 'is_active', 'trial_days')
    fieldsets = (
        ('Basic Information', {
            'fields': ('name', 'slug', 'description', 'plan_type', 'display_order', 'is_active', 'is_recommended')
        }),
        ('Pricing', {
            'fields': ('price_monthly', 'price_yearly', 'currency', 'trial_days')
        }),
        ('Stripe Integration', {
            'fields': ('stripe_product_id', 'stripe_price_id_monthly', 'stripe_price_id_yearly'),
            'classes': ('collapse',)
        }),
        ('Metadata', {
            'fields': ('metadata',),
            'classes': ('collapse',)
        }),
        ('Audit', {
            'fields': ('id', 'created_at', 'updated_at', 'created_by', 'updated_by'),
            'classes': ('collapse',)
        }),
    )
    def get_queryset(self, request):
        return super().get_queryset(request).prefetch_related('features')
    def feature_count(self, obj):
        return obj.features.count()
    feature_count.short_description = 'Features'

@admin.register(PlanFeature)
class PlanFeatureAdmin(BaseBillingAdmin):
    list_display = ('plan', 'name', 'value', 'is_highlight', 'display_order')
    list_filter = ('plan', 'is_highlight')
    search_fields = ('name', 'value')
    list_editable = ('value', 'is_highlight', 'display_order')

@admin.register(Price)
class PriceAdmin(BaseBillingAdmin):
    list_display = ('stripe_price_id', 'plan', 'amount', 'currency', 'interval', 'interval_count', 'is_active')
    list_filter = ('currency', 'interval', 'is_active')
    search_fields = ('stripe_price_id', 'stripe_product_id', 'plan__name')
    readonly_fields = ('last_synced_at',)

@admin.register(Subscription)
class SubscriptionAdmin(BaseBillingAdmin):
    list_display = ('tenant_link', 'plan_link', 'status', 'billing_interval', 'current_period_end', 'is_active_display')
    list_filter = ('status', 'billing_interval', 'plan__plan_type')
    search_fields = ('tenant__name', 'tenant__slug', 'stripe_customer_id', 'stripe_subscription_id')
    raw_id_fields = ('tenant', 'plan')
    readonly_fields = ('stripe_customer_id', 'stripe_subscription_id', 'features_snapshot')
    fieldsets = (
        ('Tenant & Plan', {
            'fields': ('tenant', 'plan', 'status', 'billing_interval')
        }),
        ('Stripe Integration', {
            'fields': ('stripe_customer_id', 'stripe_subscription_id', 'stripe_price_id'),
            'classes': ('collapse',)
        }),
        ('Dates', {
            'fields': ('trial_start', 'trial_end', 'current_period_start', 'current_period_end', 
                      'canceled_at', 'ended_at', 'cancel_at_period_end')
        }),
        ('Billing', {
            'fields': ('auto_renew', 'features_snapshot')
        }),
        ('Audit', {
            'fields': ('id', 'created_at', 'updated_at', 'created_by', 'updated_by'),
            'classes': ('collapse',)
        }),
    )
    def tenant_link(self, obj):
        from django.contrib.contenttypes.models import ContentType
        url = reverse('admin:tenant_client_change', args=[obj.tenant.id])
        return format_html('<a href="{}">{}</a>', url, obj.tenant.name)
    tenant_link.short_description = 'Tenant'
    def plan_link(self, obj):
        url = reverse('admin:billing_plan_change', args=[obj.plan.id])
        return format_html('<a href="{}">{}</a>', url, obj.plan.name)
    plan_link.short_description = 'Plan'
    def is_active_display(self, obj):
        if obj.is_active:
            return format_html('<span style="color: green;">✓ Active</span>')
        return format_html('<span style="color: red;">✗ Inactive</span>')
    is_active_display.short_description = 'Active'
    actions = ['sync_with_stripe', 'send_invoice_reminder']
    def sync_with_stripe(self, request, queryset):
        from billing.tasks import sync_subscription_with_stripe
        count = 0
        for sub in queryset:
            if sub.stripe_subscription_id:
                sync_subscription_with_stripe.delay(str(sub.id))
                count += 1
        self.message_user(request, f'Queued {count} subscriptions for sync.')
    sync_with_stripe.short_description = 'Sync selected with Stripe'
    def send_invoice_reminder(self, request, queryset):
        from billing.tasks import send_upcoming_invoice_reminder
        count = 0
        for sub in queryset:
            send_upcoming_invoice_reminder.delay(str(sub.id))
            count += 1
        self.message_user(request, f'Sent reminders to {count} subscriptions.')
    send_invoice_reminder.short_description = 'Send invoice reminders'

@admin.register(SubscriptionHistory)
class SubscriptionHistoryAdmin(BaseBillingAdmin):
    list_display = ('subscription', 'previous_status', 'new_status', 'change_reason', 'created_at')
    list_filter = ('previous_status', 'new_status')
    search_fields = ('subscription__tenant__name', 'change_reason')
    readonly_fields = ('subscription', 'previous_plan', 'new_plan', 'previous_status', 'new_status', 'change_reason', 'metadata')
    def has_add_permission(self, request):
        return False
    def has_change_permission(self, request, obj=None):
        return False

@admin.register(Invoice)
class InvoiceAdmin(BaseBillingAdmin):
    list_display = ('invoice_number', 'tenant_link', 'status', 'amount_due', 'currency', 'invoice_date', 'due_date', 'is_overdue_display')
    list_filter = ('status', 'currency', 'invoice_date')
    search_fields = ('invoice_number', 'tenant__name', 'stripe_invoice_id')
    raw_id_fields = ('tenant', 'subscription')
    readonly_fields = ('stripe_invoice_id', 'stripe_payment_intent_id', 'invoice_pdf_url')
    fieldsets = (
        ('Invoice Details', {
            'fields': ('invoice_number', 'tenant', 'subscription', 'status')
        }),
        ('Amounts', {
            'fields': ('amount_due', 'amount_paid', 'amount_remaining', 'currency')
        }),
        ('Dates', {
            'fields': ('invoice_date', 'due_date', 'period_start', 'period_end')
        }),
        ('Stripe Integration', {
            'fields': ('stripe_invoice_id', 'stripe_payment_intent_id', 'invoice_pdf_url'),
            'classes': ('collapse',)
        }),
        ('Metadata', {
            'fields': ('metadata',),
            'classes': ('collapse',)
        }),
        ('Audit', {
            'fields': ('id', 'created_at', 'updated_at', 'created_by', 'updated_by'),
            'classes': ('collapse',)
        }),
    )
    def tenant_link(self, obj):
        url = reverse('admin:tenant_client_change', args=[obj.tenant.id])
        return format_html('<a href="{}">{}</a>', url, obj.tenant.name)
    tenant_link.short_description = 'Tenant'
    def is_overdue_display(self, obj):
        if obj.is_overdue:
            return format_html('<span style="color: red;">⚠ Overdue</span>')
        return format_html('<span style="color: green;">✓ Current</span>')
    is_overdue_display.short_description = 'Status'
    actions = ['mark_as_paid', 'send_reminder']
    def mark_as_paid(self, request, queryset):
        updated = queryset.update(status='paid')
        self.message_user(request, f'Marked {updated} invoices as paid.')
    mark_as_paid.short_description = 'Mark selected as paid'
    def send_reminder(self, request, queryset):
        from billing.tasks import send_invoice_reminder_email
        count = 0
        for invoice in queryset:
            send_invoice_reminder_email.delay(str(invoice.id))
            count += 1
        self.message_user(request, f'Sent reminders for {count} invoices.')
    send_reminder.short_description = 'Send payment reminders'

@admin.register(InvoiceLineItem)
class InvoiceLineItemAdmin(BaseBillingAdmin):
    list_display = ('invoice', 'description', 'quantity', 'unit_amount', 'amount')
    list_filter = ('invoice__status',)
    search_fields = ('description', 'invoice__invoice_number')
    readonly_fields = ('stripe_price_id',)
    def has_add_permission(self, request):
        return False

@admin.register(Payment)
class PaymentAdmin(BaseBillingAdmin):
    list_display = ('id', 'tenant_link', 'amount', 'currency', 'status', 'payment_date', 'receipt_link')
    list_filter = ('status', 'currency', 'payment_date')
    search_fields = ('stripe_payment_intent_id', 'stripe_charge_id', 'tenant__name')
    raw_id_fields = ('tenant', 'subscription', 'invoice', 'payment_method')
    readonly_fields = ('stripe_payment_intent_id', 'stripe_charge_id', 'receipt_url')
    fieldsets = (
        ('Payment Details', {
            'fields': ('tenant', 'subscription', 'invoice', 'payment_method', 'status')
        }),
        ('Amounts', {
            'fields': ('amount', 'currency', 'refunded_amount')
        }),
        ('Dates', {
            'fields': ('payment_date', 'refunded_at')
        }),
        ('Stripe Integration', {
            'fields': ('stripe_payment_intent_id', 'stripe_charge_id', 'receipt_url'),
            'classes': ('collapse',)
        }),
        ('Failure Info', {
            'fields': ('failure_reason',),
            'classes': ('collapse',)
        }),
        ('Metadata', {
            'fields': ('metadata',),
            'classes': ('collapse',)
        }),
        ('Audit', {
            'fields': ('id', 'created_at', 'updated_at', 'created_by', 'updated_by'),
            'classes': ('collapse',)
        }),
    )
    
    def tenant_link(self, obj):
        url = reverse('admin:tenant_client_change', args=[obj.tenant.id])
        return format_html('<a href="{}">{}</a>', url, obj.tenant.name)
    tenant_link.short_description = 'Tenant'
    
    def receipt_link(self, obj):
        if obj.receipt_url:
            return format_html('<a href="{}" target="_blank">View Receipt</a>', obj.receipt_url)
        return '-'
    receipt_link.short_description = 'Receipt'
    
    def get_queryset(self, request):
        return super().get_queryset(request).select_related('tenant', 'subscription')


@admin.register(PaymentMethod)
class PaymentMethodAdmin(BaseBillingAdmin):
    list_display = ('tenant', 'method_type', 'last4', 'brand', 'is_default', 'is_active')
    list_filter = ('method_type', 'brand', 'is_default', 'is_active')
    search_fields = ('tenant__name', 'last4', 'billing_email')
    raw_id_fields = ('tenant', 'subscription')
    readonly_fields = ('stripe_payment_method_id', 'stripe_customer_id')


@admin.register(WebhookEvent)
class WebhookEventAdmin(BaseBillingAdmin):
    list_display = ('event_type', 'stripe_event_id', 'is_processed', 'retry_count', 'created_at')
    list_filter = ('event_type', 'is_processed', 'api_version')
    search_fields = ('stripe_event_id', 'event_type', 'processing_error')
    readonly_fields = ('stripe_event_id', 'event_type', 'api_version', 'payload', 'is_processed', 'processed_at', 'processing_error', 'retry_count')
    
    def has_add_permission(self, request):
        return False
    
    def has_change_permission(self, request, obj=None):
        return False
        
    actions = ['reprocess_events']
    def reprocess_events(self, request, queryset):
        from billing.tasks import process_webhook_event
        count = 0
        for event in queryset.filter(is_processed=False):
            process_webhook_event.delay(str(event.id))
            count += 1
        self.message_user(request, f'Queued {count} webhook events for reprocessing.')
    reprocess_events.short_description = 'Reprocess selected events'


@admin.register(QuotaLimit)
class QuotaLimitAdmin(BaseBillingAdmin):
    list_display = ('subscription', 'max_users', 'max_kpis', 'max_storage_mb', 'max_api_calls_per_day')
    search_fields = ('subscription__tenant__name',)
    raw_id_fields = ('subscription',)
    fieldsets = (
        ('User Limits', {
            'fields': ('max_users', 'max_admins')
        }),
        ('KPI Limits', {
            'fields': ('max_kpis', 'max_kpi_frameworks')
        }),
        ('Storage & API', {
            'fields': ('max_storage_mb', 'max_api_calls_per_day')
        }),
        ('Feature Flags', {
            'fields': ('allow_custom_branding', 'allow_api_access', 'allow_sso', 
                      'allow_advanced_analytics', 'allow_audit_logs', 'allow_reports', 'allow_export')
        }),
        ('Audit', {
            'fields': ('id', 'created_at', 'updated_at', 'created_by', 'updated_by'),
            'classes': ('collapse',)
        }),
    )

@admin.register(QuotaUsage)
class QuotaUsageAdmin(BaseBillingAdmin):
    list_display = ('tenant', 'current_users', 'current_admins', 'current_kpis', 'current_storage_mb', 'api_calls_today', 'snapshot_date')
    list_filter = ('snapshot_date',)
    search_fields = ('tenant__name',)
    raw_id_fields = ('tenant',)
    readonly_fields = ('snapshot_date',)
    
    def has_add_permission(self, request):
        return False

class ExportMixin:
    def export_to_csv(self, request, queryset, filename=None, fields=None):
        if not filename:
            filename = f"{self.model._meta.model_name}_{timezone.now().strftime('%Y%m%d_%H%M%S')}.csv"
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        writer = csv.writer(response)
        if fields is None:
            fields = [f.name for f in self.model._meta.fields if not f.is_relation]
        writer.writerow(fields)
        for obj in queryset:
            row = []
            for field in fields:
                value = getattr(obj, field, '')
                if callable(value):
                    value = value()
                elif hasattr(value, 'strftime'):
                    value = value.strftime('%Y-%m-%d %H:%M:%S')
                row.append(str(value) if value else '')
            writer.writerow(row)
        return response
    
    def export_to_excel(self, request, queryset, filename=None, fields=None):
        if not filename:
            filename = f"{self.model._meta.model_name}_{timezone.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        wb = Workbook()
        ws = wb.active
        ws.title = self.model._meta.verbose_name_plural.title()
        if fields is None:
            fields = [f.name for f in self.model._meta.fields if not f.is_relation]
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        for col_idx, field in enumerate(fields, 1):
            cell = ws.cell(row=1, column=col_idx, value=field.replace('_', ' ').title())
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        for row_idx, obj in enumerate(queryset, 2):
            for col_idx, field in enumerate(fields, 1):
                value = getattr(obj, field, '')
                if callable(value):
                    value = value()
                elif hasattr(value, 'strftime'):
                    value = value.strftime('%Y-%m-%d %H:%M:%S')
                cell = ws.cell(row=row_idx, column=col_idx, value=str(value) if value else '')
                cell.alignment = Alignment(horizontal="left", vertical="center")
        for col_idx, field in enumerate(fields, 1):
            column_letter = get_column_letter(col_idx)
            max_length = max(
                len(field),
                max([len(str(getattr(obj, field, ''))) for obj in queryset[:100]] + [0])
            )
            ws.column_dimensions[column_letter].width = min(max_length + 2, 50)
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        wb.save(response)
        return response
    
    def export_to_pdf(self, request, queryset, filename=None, title=None):
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import letter, landscape
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.units import inch
        if not filename:
            filename = f"{self.model._meta.model_name}_{timezone.now().strftime('%Y%m%d_%H%M%S')}.pdf"
        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=landscape(letter))
        styles = getSampleStyleSheet()
        elements = []
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=16,
            spaceAfter=30,
            alignment=1  # Center
        )
        title_text = title or f"{self.model._meta.verbose_name_plural.title()} Export"
        elements.append(Paragraph(title_text, title_style))
        elements.append(Spacer(1, 0.2 * inch))
        date_style = ParagraphStyle(
            'DateStyle',
            parent=styles['Normal'],
            fontSize=10,
            textColor=colors.grey,
            alignment=2  # Right
        )
        elements.append(Paragraph(f"Generated: {timezone.now().strftime('%Y-%m-%d %H:%M:%S')}", date_style))
        elements.append(Spacer(1, 0.2 * inch))
        fields = [f.name for f in self.model._meta.fields if not f.is_relation and f.name not in ['payload', 'metadata']]
        field_labels = [f.replace('_', ' ').title() for f in fields]
        data = [field_labels]
        for obj in queryset[:500]:  # Limit to 500 rows for PDF
            row = []
            for field in fields:
                value = getattr(obj, field, '')
                if callable(value):
                    value = value()
                elif hasattr(value, 'strftime'):
                    value = value.strftime('%Y-%m-%d')
                row.append(str(value)[:100] if value else '')  # Limit cell size
            data.append(row)
        table = Table(data)
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4F81BD')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.grey),
            ('FONTSIZE', (0, 1), (-1, -1), 8),
            ('TOPPADDING', (0, 1), (-1, -1), 6),
            ('BOTTOMPADDING', (0, 1), (-1, -1), 6),
        ]))
        elements.append(table)
        doc.build(elements)
        response.write(buffer.getvalue())
        buffer.close()
        return response

class EnhancedSubscriptionAdmin(ExportMixin):
    actions = ['export_subscriptions_csv', 'export_subscriptions_excel', 'export_subscriptions_pdf']
    def export_subscriptions_csv(self, request, queryset):
        return self.export_to_csv(request, queryset, 'subscriptions_export.csv')
    export_subscriptions_csv.short_description = 'Export selected to CSV'
    
    def export_subscriptions_excel(self, request, queryset):
        return self.export_to_excel(request, queryset, 'subscriptions_export.xlsx')
    export_subscriptions_excel.short_description = 'Export selected to Excel'
    
    def export_subscriptions_pdf(self, request, queryset):
        return self.export_to_pdf(request, queryset, 'subscriptions_export.pdf', 'Subscriptions Report')
    export_subscriptions_pdf.short_description = 'Export selected to PDF'

class EnhancedInvoiceAdmin(ExportMixin):
    actions = ['export_invoices_csv', 'export_invoices_excel', 'export_invoices_pdf', 'send_bulk_reminders']
    
    def export_invoices_csv(self, request, queryset):
        fields = ['invoice_number', 'tenant__name', 'amount_due', 'currency', 'status', 'invoice_date', 'due_date']
        return self.export_to_csv(request, queryset, 'invoices_export.csv', fields)
    export_invoices_csv.short_description = 'Export selected to CSV'
    
    def export_invoices_excel(self, request, queryset):
        fields = ['invoice_number', 'tenant__name', 'amount_due', 'currency', 'status', 'invoice_date', 'due_date']
        return self.export_to_excel(request, queryset, 'invoices_export.xlsx', fields)
    export_invoices_excel.short_description = 'Export selected to Excel'
    
    def export_invoices_pdf(self, request, queryset):
        return self.export_to_pdf(request, queryset, 'invoices_export.pdf', 'Invoices Report')
    export_invoices_pdf.short_description = 'Export selected to PDF'
    
    def send_bulk_reminders(self, request, queryset):
        from billing.tasks import send_invoice_reminder_email
        count = 0
        for invoice in queryset.filter(status__in=['draft', 'open']):
            send_invoice_reminder_email.delay(str(invoice.id))
            count += 1
        self.message_user(request, f'Sent reminders for {count} invoices.')
    send_bulk_reminders.short_description = 'Send payment reminders to selected'

class EnhancedPaymentAdmin(ExportMixin):
    actions = ['export_payments_csv', 'export_payments_excel', 'export_payments_pdf']
    def export_payments_csv(self, request, queryset):
        fields = ['tenant__name', 'amount', 'currency', 'status', 'payment_date', 'failure_reason']
        return self.export_to_csv(request, queryset, 'payments_export.csv', fields)
    export_payments_csv.short_description = 'Export selected to CSV'
    
    def export_payments_excel(self, request, queryset):
        fields = ['tenant__name', 'amount', 'currency', 'status', 'payment_date', 'failure_reason']
        return self.export_to_excel(request, queryset, 'payments_export.xlsx', fields)
    export_payments_excel.short_description = 'Export selected to Excel'
    
    def export_payments_pdf(self, request, queryset):
        return self.export_to_pdf(request, queryset, 'payments_export.pdf', 'Payments Report')
    export_payments_pdf.short_description = 'Export selected to PDF'