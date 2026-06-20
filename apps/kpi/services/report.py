# report.py
import csv
import io
import logging
from datetime import datetime
from decimal import Decimal
from typing import List, Dict, Optional, Any
from django.db.models import Q, Avg, Count
from django.utils import timezone
from django.core.cache import cache
from django.core.exceptions import PermissionDenied, ValidationError
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from ..models import KPI, Score, KPISummary, TrafficLight, MonthlyActual
from .analytics import get_department_rollups, get_organization_health, get_kpi_summaries

logger = logging.getLogger(__name__)

CACHE_TTL = 3600
CACHE_PREFIX = "kpi_report"


class ReportGenerator:
    def __init__(self):
        self.styles = getSampleStyleSheet()

    def _check_tenant_access(self, tenant_id: str, user_tenant_id: str) -> None:
        if str(tenant_id) != str(user_tenant_id):
            raise PermissionDenied("Access denied to this tenant's data")

    def generate_pdf_report(
        self,
        tenant_id: str,
        user_tenant_id: str,
        year: int = None,
        month: int = None
    ) -> bytes:
        self._check_tenant_access(tenant_id, user_tenant_id)

        if not year:
            year = timezone.now().year
        if not month:
            month = timezone.now().month

        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4, title="KPI Performance Report")
        story = []

        title_style = ParagraphStyle(
            'CustomTitle',
            parent=self.styles['Heading1'],
            fontSize=16,
            spaceAfter=30,
            alignment=1
        )
        title = Paragraph("KPI Performance Report", title_style)
        story.append(title)

        date_str = f"{datetime(year, month, 1).strftime('%B %Y')}"
        date_para = Paragraph(f"Report Period: {date_str}", self.styles['Normal'])
        story.append(date_para)
        story.append(Spacer(1, 20))

        self._add_executive_summary(story, tenant_id, year, month)
        self._add_kpi_performance_table(story, tenant_id, year, month)
        self._add_department_summary(story, tenant_id, year, month)
        self._add_red_alerts_table(story, tenant_id, year, month)

        doc.build(story)
        buffer.seek(0)
        return buffer.getvalue()

    def generate_excel_report(
        self,
        tenant_id: str,
        user_tenant_id: str,
        year: int = None,
        month: int = None
    ) -> bytes:
        self._check_tenant_access(tenant_id, user_tenant_id)

        if not year:
            year = timezone.now().year
        if not month:
            month = timezone.now().month

        buffer = io.BytesIO()
        wb = Workbook()

        ws_summary = wb.active
        ws_summary.title = "Executive Summary"

        ws_kpis = wb.create_sheet("KPI Details")
        ws_departments = wb.create_sheet("Department Details")
        ws_alerts = wb.create_sheet("Red Alerts")

        self._add_excel_executive_summary(ws_summary, tenant_id, year, month)
        self._add_excel_kpi_details(ws_kpis, tenant_id, year, month)
        self._add_excel_department_details(ws_departments, tenant_id, year, month)
        self._add_excel_red_alerts(ws_alerts, tenant_id, year, month)

        wb.save(buffer)
        buffer.seek(0)
        return buffer.getvalue()

    def generate_kpi_performance_report(
        self,
        tenant_id: str,
        user_tenant_id: str,
        kpi_ids: List[str] = None,
        year: int = None,
        month: int = None,
        format_type: str = 'pdf'
    ) -> Dict:
        self._check_tenant_access(tenant_id, user_tenant_id)

        if not year:
            year = timezone.now().year
        if not month:
            month = timezone.now().month

        cache_key = f"{CACHE_PREFIX}:kpi_perf:{tenant_id}:{year}:{month}:{hash(str(kpi_ids))}"
        cached = cache.get(cache_key)
        if cached:
            return cached

        kpis = KPI.objects.filter(tenant_id=tenant_id)
        if kpi_ids:
            kpis = kpis.filter(id__in=kpi_ids)

        data = []
        for kpi in kpis:
            scores = Score.objects.filter(kpi=kpi, year=year, month=month)
            avg_score = scores.aggregate(avg=Avg('score'))['avg'] or 0

            data.append({
                'kpi_id': str(kpi.id),
                'kpi_name': kpi.name,
                'kpi_code': kpi.code,
                'average_score': float(avg_score),
                'green_count': scores.filter(score__gte=90).count(),
                'yellow_count': scores.filter(score__gte=50, score__lt=90).count(),
                'red_count': scores.filter(score__lt=50).count(),
                'total_users': scores.count()
            })

        if format_type == 'pdf':
            result = {'type': 'pdf', 'data': self._generate_kpi_pdf_report(data, year, month)}
        elif format_type == 'excel':
            result = {'type': 'excel', 'data': self._generate_kpi_excel_report(data, year, month)}
        else:
            result = {'type': 'csv', 'data': self._generate_kpi_csv_report(data, year, month)}

        cache.set(cache_key, result, CACHE_TTL)
        return result

    def generate_department_comparison_report(
        self,
        tenant_id: str,
        user_tenant_id: str,
        year: int = None,
        month: int = None,
        format_type: str = 'pdf'
    ) -> Dict:
        self._check_tenant_access(tenant_id, user_tenant_id)

        if not year:
            year = timezone.now().year
        if not month:
            month = timezone.now().month

        cache_key = f"{CACHE_PREFIX}:dept_comp:{tenant_id}:{year}:{month}"
        cached = cache.get(cache_key)
        if cached:
            return cached

        rollups = get_department_rollups(str(tenant_id), year, month)

        if format_type == 'pdf':
            result = {'type': 'pdf', 'data': self._generate_department_pdf_report(rollups, year, month)}
        elif format_type == 'excel':
            result = {'type': 'excel', 'data': self._generate_department_excel_report(rollups, year, month)}
        else:
            result = {'type': 'csv', 'data': self._generate_department_csv_report(rollups, year, month)}

        cache.set(cache_key, result, CACHE_TTL)
        return result

    def generate_trend_analysis_report(
        self,
        tenant_id: str,
        user_tenant_id: str,
        kpi_ids: List[str] = None,
        months: int = 12,
        format_type: str = 'pdf'
    ) -> Dict:
        self._check_tenant_access(tenant_id, user_tenant_id)

        end_date = timezone.now()
        start_date = end_date - timezone.timedelta(days=30 * months)

        kpis = KPI.objects.filter(tenant_id=tenant_id)
        if kpi_ids:
            kpis = kpis.filter(id__in=kpi_ids)

        trends = []
        for kpi in kpis:
            scores = Score.objects.filter(
                kpi=kpi,
                calculated_at__gte=start_date,
                calculated_at__lte=end_date
            ).order_by('year', 'month')

            monthly_scores = []
            for score in scores:
                monthly_scores.append({
                    'period': f"{score.year}-{score.month:02d}",
                    'score': float(score.score)
                })

            avg_score = sum(s['score'] for s in monthly_scores) / len(monthly_scores) if monthly_scores else 0

            trends.append({
                'kpi_id': str(kpi.id),
                'kpi_name': kpi.name,
                'scores': monthly_scores,
                'average': round(avg_score, 2),
                'trend': self._calculate_trend([s['score'] for s in monthly_scores])
            })

        if format_type == 'pdf':
            return {'type': 'pdf', 'data': self._generate_trend_pdf_report(trends, months)}
        else:
            return {'type': 'csv', 'data': self._generate_trend_csv_report(trends, months)}

    def generate_monthly_report(self, tenant_id: str, year: int, month: int) -> Dict:
        health = get_organization_health(str(tenant_id), year, month)

        return {
            'id': f"monthly_{year}_{month:02d}",
            'tenant_id': tenant_id,
            'period': f"{year}-{month:02d}",
            'generated_at': timezone.now().isoformat(),
            'organization_health': health,
            'status': 'COMPLETED'
        }

    def generate_quarterly_report(self, tenant_id: str, year: int, quarter: int) -> Dict:
        months = {
            1: [1, 2, 3],
            2: [4, 5, 6],
            3: [7, 8, 9],
            4: [10, 11, 12]
        }.get(quarter, [1, 2, 3])

        monthly_reports = []
        for month in months:
            report = self.generate_monthly_report(tenant_id, year, month)
            monthly_reports.append(report)

        avg_health = sum(
            r['organization_health'].get('overall_health_score', 0)
            for r in monthly_reports
        ) / len(monthly_reports) if monthly_reports else 0

        return {
            'id': f"quarterly_{year}_Q{quarter}",
            'tenant_id': tenant_id,
            'period': f"Q{quarter} {year}",
            'generated_at': timezone.now().isoformat(),
            'average_health': round(avg_health, 2),
            'monthly_reports': monthly_reports,
            'status': 'COMPLETED'
        }

    def generate_annual_report(self, tenant_id: str, year: int) -> Dict:
        quarterly_reports = []
        for q in range(1, 5):
            report = self.generate_quarterly_report(tenant_id, year, q)
            quarterly_reports.append(report)

        avg_health = sum(q['average_health'] for q in quarterly_reports) / len(quarterly_reports) if quarterly_reports else 0

        return {
            'id': f"annual_{year}",
            'tenant_id': tenant_id,
            'period': str(year),
            'generated_at': timezone.now().isoformat(),
            'average_health': round(avg_health, 2),
            'quarterly_reports': quarterly_reports,
            'status': 'COMPLETED'
        }

    def generate_custom_report(
        self,
        tenant_id: str,
        report_data: Dict,
        user_tenant_id: str
    ) -> Dict:
        self._check_tenant_access(tenant_id, user_tenant_id)

        report_type = report_data.get('report_type')
        filters = report_data.get('filters', {})
        format_type = report_data.get('format', 'pdf')

        if report_type == 'kpi_performance':
            return self.generate_kpi_performance_report(
                tenant_id,
                user_tenant_id,
                filters.get('kpi_ids', []),
                filters.get('year'),
                filters.get('month'),
                format_type
            )
        elif report_type == 'department_comparison':
            return self.generate_department_comparison_report(
                tenant_id,
                user_tenant_id,
                filters.get('year'),
                filters.get('month'),
                format_type
            )
        elif report_type == 'trend_analysis':
            return self.generate_trend_analysis_report(
                tenant_id,
                user_tenant_id,
                filters.get('kpi_ids', []),
                filters.get('months', 12),
                format_type
            )
        else:
            raise ValidationError(f"Unknown report type: {report_type}")

    def _add_executive_summary(self, story, tenant_id, year, month):
        heading = Paragraph("Executive Summary", self.styles['Heading2'])
        story.append(heading)

        health = get_organization_health(str(tenant_id), year, month)

        if health:
            summary_text = f"""
            <b>Overall Health Score:</b> {health.get('overall_health_score', 0)}%<br/>
            <b>KPI Completion Rate:</b> {health.get('kpi_completion_rate', 0)}%<br/>
            <b>Validation Compliance:</b> {health.get('validation_compliance_rate', 0)}%<br/>
            <b>Red KPIs:</b> {health.get('red_kpi_count', 0)}<br/>
            <b>Active Employees:</b> {health.get('active_employees', 0)}
            """
            summary = Paragraph(summary_text, self.styles['Normal'])
            story.append(summary)
        else:
            story.append(Paragraph("No health data available.", self.styles['Normal']))

        story.append(Spacer(1, 20))

    def _add_kpi_performance_table(self, story, tenant_id, year, month):
        heading = Paragraph("KPI Performance Details", self.styles['Heading2'])
        story.append(heading)

        summaries = get_kpi_summaries(tenant_id, year, month, prefer_mv=True)

        if summaries:
            data = [['KPI Name', 'Avg Score', 'Green', 'Yellow', 'Red', 'Users']]
            for summary in summaries[:20]:
                data.append([
                    summary.get('kpi_name', 'Unknown')[:40],
                    f"{summary.get('average_score', 0):.1f}%",
                    summary.get('green_count', 0),
                    summary.get('yellow_count', 0),
                    summary.get('red_count', 0),
                    summary.get('total_users', 0)
                ])

            table = Table(data)
            table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 10),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.black),
                ('FONTSIZE', (0, 1), (-1, -1), 8),
            ]))
            story.append(table)
        else:
            story.append(Paragraph("No KPI performance data available.", self.styles['Normal']))

        story.append(Spacer(1, 20))

    def _add_department_summary(self, story, tenant_id, year, month):
        heading = Paragraph("Department Performance Summary", self.styles['Heading2'])
        story.append(heading)

        rollups = get_department_rollups(str(tenant_id), year, month)

        if rollups:
            data = [['Department', 'Score', 'Green %', 'Yellow %', 'Red %', 'Employees']]
            for rollup in rollups[:15]:
                data.append([
                    rollup.get('department_name', 'Unknown')[:30],
                    f"{rollup.get('overall_score', 0):.1f}%",
                    f"{rollup.get('green_percentage', 0):.1f}%",
                    f"{rollup.get('yellow_percentage', 0):.1f}%",
                    f"{rollup.get('red_percentage', 0):.1f}%",
                    rollup.get('employee_count', 0)
                ])

            table = Table(data)
            table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.black),
            ]))
            story.append(table)
        else:
            story.append(Paragraph("No department performance data available.", self.styles['Normal']))

        story.append(Spacer(1, 20))

    def _add_red_alerts_table(self, story, tenant_id, year, month):
        heading = Paragraph("Red Alerts (2+ Consecutive Red Months)", self.styles['Heading2'])
        story.append(heading)

        red_alerts = TrafficLight.objects.filter(
            score__tenant_id=tenant_id,
            score__year=year,
            score__month=month,
            status='RED',
            consecutive_red_count__gte=2
        ).select_related('score__kpi', 'score__user')[:50]

        if red_alerts.exists():
            data = [['KPI', 'User', 'Score', 'Consecutive Months']]
            for alert in red_alerts:
                data.append([
                    alert.score.kpi.name[:40],
                    alert.score.user.email,
                    f"{alert.score_value:.1f}%",
                    alert.consecutive_red_count
                ])

            table = Table(data)
            table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.red),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.black),
            ]))
            story.append(table)
        else:
            story.append(Paragraph("No red alerts for this period.", self.styles['Normal']))

    def _add_excel_executive_summary(self, ws, tenant_id, year, month):
        ws['A1'] = "KPI Performance Report"
        ws['A1'].font = Font(size=16, bold=True)

        ws['A2'] = f"Report Period: {datetime(year, month, 1).strftime('%B %Y')}"
        ws['A2'].font = Font(size=12)

        ws['A4'] = "Executive Summary"
        ws['A4'].font = Font(size=14, bold=True)

        health = get_organization_health(str(tenant_id), year, month)

        if health:
            ws['A5'] = f"Overall Health Score: {health.get('overall_health_score', 0)}%"
            ws['A6'] = f"KPI Completion Rate: {health.get('kpi_completion_rate', 0)}%"
            ws['A7'] = f"Validation Compliance: {health.get('validation_compliance_rate', 0)}%"
            ws['A8'] = f"Red KPIs: {health.get('red_kpi_count', 0)}"
            ws['A9'] = f"Active Employees: {health.get('active_employees', 0)}"

    def _add_excel_kpi_details(self, ws, tenant_id, year, month):
        ws['A1'] = "KPI Performance Details"
        ws['A1'].font = Font(size=14, bold=True)

        headers = ['KPI Name', 'Average Score', 'Green', 'Yellow', 'Red', 'Total Users']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=2, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
            cell.alignment = Alignment(horizontal='center')

        summaries = get_kpi_summaries(tenant_id, year, month, prefer_mv=True)

        for row, summary in enumerate(summaries, 3):
            ws.cell(row=row, column=1, value=summary.get('kpi_name', 'Unknown'))
            ws.cell(row=row, column=2, value=f"{summary.get('average_score', 0):.1f}%")
            ws.cell(row=row, column=3, value=summary.get('green_count', 0))
            ws.cell(row=row, column=4, value=summary.get('yellow_count', 0))
            ws.cell(row=row, column=5, value=summary.get('red_count', 0))
            ws.cell(row=row, column=6, value=summary.get('total_users', 0))

        for col in range(1, 7):
            ws.column_dimensions[chr(64 + col)].width = 20

    def _add_excel_department_details(self, ws, tenant_id, year, month):
        ws['A1'] = "Department Performance Summary"
        ws['A1'].font = Font(size=14, bold=True)

        headers = ['Department', 'Overall Score', 'Green %', 'Yellow %', 'Red %', 'Employees']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=2, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
            cell.alignment = Alignment(horizontal='center')

        rollups = get_department_rollups(str(tenant_id), year, month)

        for row, rollup in enumerate(rollups, 3):
            ws.cell(row=row, column=1, value=rollup.get('department_name', 'Unknown'))
            ws.cell(row=row, column=2, value=f"{rollup.get('overall_score', 0):.1f}%")
            ws.cell(row=row, column=3, value=f"{rollup.get('green_percentage', 0):.1f}%")
            ws.cell(row=row, column=4, value=f"{rollup.get('yellow_percentage', 0):.1f}%")
            ws.cell(row=row, column=5, value=f"{rollup.get('red_percentage', 0):.1f}%")
            ws.cell(row=row, column=6, value=rollup.get('employee_count', 0))

    def _add_excel_red_alerts(self, ws, tenant_id, year, month):
        ws['A1'] = "Red Alerts (2+ Consecutive Red Months)"
        ws['A1'].font = Font(size=14, bold=True)

        headers = ['KPI', 'User', 'Score', 'Consecutive Months']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=2, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="FF9999", end_color="FF9999", fill_type="solid")
            cell.alignment = Alignment(horizontal='center')

        red_alerts = TrafficLight.objects.filter(
            score__tenant_id=tenant_id,
            score__year=year,
            score__month=month,
            status='RED',
            consecutive_red_count__gte=2
        ).select_related('score__kpi', 'score__user')[:100]

        for row, alert in enumerate(red_alerts, 3):
            ws.cell(row=row, column=1, value=alert.score.kpi.name)
            ws.cell(row=row, column=2, value=alert.score.user.email)
            ws.cell(row=row, column=3, value=f"{alert.score_value:.1f}%")
            ws.cell(row=row, column=4, value=alert.consecutive_red_count)

    def _generate_kpi_pdf_report(self, data: List[Dict], year: int, month: int) -> bytes:
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4)
        story = []

        title = Paragraph(
            f"KPI Performance Report - {datetime(year, month, 1).strftime('%B %Y')}",
            ParagraphStyle('Title', parent=self.styles['Heading1'], fontSize=14)
        )
        story.append(title)
        story.append(Spacer(1, 20))

        if data:
            table_data = [['KPI Name', 'Avg Score', 'Green', 'Yellow', 'Red', 'Users']]
            for item in data:
                table_data.append([
                    item['kpi_name'][:40],
                    f"{item['average_score']:.1f}%",
                    item['green_count'],
                    item['yellow_count'],
                    item['red_count'],
                    item['total_users']
                ])

            table = Table(table_data)
            table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.black),
            ]))
            story.append(table)
        else:
            story.append(Paragraph("No data available", self.styles['Normal']))

        doc.build(story)
        buffer.seek(0)
        return buffer.getvalue()

    def _generate_kpi_excel_report(self, data: List[Dict], year: int, month: int) -> bytes:
        buffer = io.BytesIO()
        wb = Workbook()
        ws = wb.active
        ws.title = "KPI Performance"

        ws['A1'] = f"KPI Performance Report - {datetime(year, month, 1).strftime('%B %Y')}"
        ws['A1'].font = Font(size=14, bold=True)

        headers = ['KPI Name', 'Average Score', 'Green', 'Yellow', 'Red', 'Total Users']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")

        for row, item in enumerate(data, 4):
            ws.cell(row=row, column=1, value=item['kpi_name'])
            ws.cell(row=row, column=2, value=f"{item['average_score']:.1f}%")
            ws.cell(row=row, column=3, value=item['green_count'])
            ws.cell(row=row, column=4, value=item['yellow_count'])
            ws.cell(row=row, column=5, value=item['red_count'])
            ws.cell(row=row, column=6, value=item['total_users'])

        wb.save(buffer)
        buffer.seek(0)
        return buffer.getvalue()

    def _generate_kpi_csv_report(self, data: List[Dict], year: int, month: int) -> str:
        output = io.StringIO()
        writer = csv.writer(output)
        writer.writerow(['KPI Name', 'Average Score', 'Green', 'Yellow', 'Red', 'Total Users'])
        for item in data:
            writer.writerow([
                item['kpi_name'],
                f"{item['average_score']:.1f}%",
                item['green_count'],
                item['yellow_count'],
                item['red_count'],
                item['total_users']
            ])
        return output.getvalue()

    def _generate_department_pdf_report(self, rollups: List[Dict], year: int, month: int) -> bytes:
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=landscape(A4))
        story = []

        title = Paragraph(
            f"Department Comparison Report - {datetime(year, month, 1).strftime('%B %Y')}",
            ParagraphStyle('Title', parent=self.styles['Heading1'], fontSize=14)
        )
        story.append(title)
        story.append(Spacer(1, 20))

        if rollups:
            table_data = [['Department', 'Score', 'Green %', 'Yellow %', 'Red %', 'Employees']]
            for rollup in rollups:
                table_data.append([
                    rollup.get('department_name', 'Unknown')[:30],
                    f"{rollup.get('overall_score', 0):.1f}%",
                    f"{rollup.get('green_percentage', 0):.1f}%",
                    f"{rollup.get('yellow_percentage', 0):.1f}%",
                    f"{rollup.get('red_percentage', 0):.1f}%",
                    rollup.get('employee_count', 0)
                ])

            table = Table(table_data)
            table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.black),
            ]))
            story.append(table)
        else:
            story.append(Paragraph("No department data available", self.styles['Normal']))

        doc.build(story)
        buffer.seek(0)
        return buffer.getvalue()

    def _generate_department_excel_report(self, rollups: List[Dict], year: int, month: int) -> bytes:
        buffer = io.BytesIO()
        wb = Workbook()
        ws = wb.active
        ws.title = "Department Comparison"

        ws['A1'] = f"Department Comparison Report - {datetime(year, month, 1).strftime('%B %Y')}"
        ws['A1'].font = Font(size=14, bold=True)

        headers = ['Department', 'Overall Score', 'Green %', 'Yellow %', 'Red %', 'Employees']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")

        for row, rollup in enumerate(rollups, 4):
            ws.cell(row=row, column=1, value=rollup.get('department_name', 'Unknown'))
            ws.cell(row=row, column=2, value=f"{rollup.get('overall_score', 0):.1f}%")
            ws.cell(row=row, column=3, value=f"{rollup.get('green_percentage', 0):.1f}%")
            ws.cell(row=row, column=4, value=f"{rollup.get('yellow_percentage', 0):.1f}%")
            ws.cell(row=row, column=5, value=f"{rollup.get('red_percentage', 0):.1f}%")
            ws.cell(row=row, column=6, value=rollup.get('employee_count', 0))

        wb.save(buffer)
        buffer.seek(0)
        return buffer.getvalue()

    def _generate_department_csv_report(self, rollups: List[Dict], year: int, month: int) -> str:
        output = io.StringIO()
        writer = csv.writer(output)
        writer.writerow(['Department', 'Overall Score', 'Green %', 'Yellow %', 'Red %', 'Employees'])
        for rollup in rollups:
            writer.writerow([
                rollup.get('department_name', 'Unknown'),
                f"{rollup.get('overall_score', 0):.1f}%",
                f"{rollup.get('green_percentage', 0):.1f}%",
                f"{rollup.get('yellow_percentage', 0):.1f}%",
                f"{rollup.get('red_percentage', 0):.1f}%",
                rollup.get('employee_count', 0)
            ])
        return output.getvalue()

    def _generate_trend_pdf_report(self, trends: List[Dict], months: int) -> bytes:
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4)
        story = []

        title = Paragraph(
            f"Trend Analysis Report - Last {months} Months",
            ParagraphStyle('Title', parent=self.styles['Heading1'], fontSize=14)
        )
        story.append(title)
        story.append(Spacer(1, 20))

        for trend in trends:
            story.append(Paragraph(f"<b>{trend['kpi_name']}</b>", self.styles['Normal']))
            story.append(Paragraph(
                f"Average Score: {trend['average']:.1f}% | Trend: {trend['trend']['direction']}",
                self.styles['Normal']
            ))
            story.append(Spacer(1, 10))

        doc.build(story)
        buffer.seek(0)
        return buffer.getvalue()

    def _generate_trend_csv_report(self, trends: List[Dict], months: int) -> str:
        output = io.StringIO()
        writer = csv.writer(output)
        writer.writerow(['KPI Name', 'Average Score', 'Trend Direction', 'Trend Confidence'])

        for trend in trends:
            writer.writerow([
                trend['kpi_name'],
                f"{trend['average']:.1f}%",
                trend['trend']['direction'],
                f"{trend['trend']['confidence']:.2f}"
            ])
        return output.getvalue()

    def _calculate_trend(self, scores: List[float]) -> Dict:
        if len(scores) < 2:
            return {'direction': 'STABLE', 'confidence': 0.0}

        first = scores[0]
        last = scores[-1]
        change = ((last - first) / first) * 100 if first > 0 else 0

        if change > 5:
            direction = 'IMPROVING'
        elif change < -5:
            direction = 'DECLINING'
        else:
            direction = 'STABLE'

        confidence = min(0.9, abs(change) / 20)

        return {'direction': direction, 'confidence': round(confidence, 2)}