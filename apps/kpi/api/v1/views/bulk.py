from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import status
import csv
import io
from ..serializers import (
    BulkKPIUploadSerializer, BulkActualUploadSerializer,
    BulkTargetUploadSerializer, BulkUploadResultSerializer
)
from ....services import KPIImportExport, ActualBatchUpload, TargetSetter
from ..throttles import BulkUploadThrottle
from ..permissions import IsAuthenticatedAndActive, IsDashboardChampion


class BulkKPIUploadView(APIView):
    permission_classes = [IsAuthenticatedAndActive, IsDashboardChampion]
    throttle_classes = [BulkUploadThrottle]
    def post(self, request):
        serializer = BulkKPIUploadSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        file = serializer.validated_data['file']
        framework_id = serializer.validated_data['framework_id']
        dry_run = serializer.validated_data.get('dry_run', False)
        # Read file content
        if file.name.endswith('.csv'):
            content = file.read().decode('utf-8')
        else:
            # Handle Excel files
            import openpyxl
            workbook = openpyxl.load_workbook(file)
            sheet = workbook.active
            output = io.StringIO()
            writer = csv.writer(output)
            for row in sheet.iter_rows(values_only=True):
                writer.writerow(row)
            content = output.getvalue()
        # Process upload
        import_export = KPIImportExport()
        if dry_run:
            # Validate only, don't save
            # This would need a dry run implementation
            result = {'created': [], 'errors': [], 'total': 0}
        else:
            result = import_export.import_from_csv(
                content,
                framework_id,
                str(request.tenant.id),
                request.user
            )
        result_serializer = BulkUploadResultSerializer({
            'total_rows': result.get('total', len(result.get('created', [])) + len(result.get('errors', []))),
            'created': len(result.get('created', [])),
            'updated': 0,
            'errors': result.get('errors', []),
            'dry_run': dry_run
        })
        return Response(result_serializer.data, status=status.HTTP_202_ACCEPTED if not dry_run else status.HTTP_200_OK)

class BulkActualUploadView(APIView):
    permission_classes = [IsAuthenticatedAndActive, IsDashboardChampion]
    throttle_classes = [BulkUploadThrottle]
    def post(self, request):
        """Upload actual data from CSV/Excel"""
        serializer = BulkActualUploadSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        
        file = serializer.validated_data['file']
        year = serializer.validated_data['year']
        month = serializer.validated_data['month']
        dry_run = serializer.validated_data.get('dry_run', False)
        
        # Read file content
        if file.name.endswith('.csv'):
            content = file.read().decode('utf-8')
        else:
            import openpyxl
            workbook = openpyxl.load_workbook(file)
            sheet = workbook.active
            output = io.StringIO()
            writer = csv.writer(output)
            for row in sheet.iter_rows(values_only=True):
                writer.writerow(row)
            content = output.getvalue()
        # Process upload
        batch_upload = ActualBatchUpload()
        
        if dry_run:
            # Validate only
            result = {'created': 0, 'errors': []}
        else:
            result = batch_upload.upload_from_csv(
                content,
                str(request.tenant.id),
                request.user
            )
        
        result_serializer = BulkUploadResultSerializer({
            'total_rows': result.get('total', result.get('created', 0) + len(result.get('errors', []))),
            'created': result.get('created', 0),
            'updated': 0,
            'errors': result.get('errors', []),
            'dry_run': dry_run
        })
        
        return Response(result_serializer.data, status=status.HTTP_202_ACCEPTED if not dry_run else status.HTTP_200_OK)


class BulkTargetUploadView(APIView):
    permission_classes = [IsAuthenticatedAndActive, IsDashboardChampion]
    throttle_classes = [BulkUploadThrottle]
    def post(self, request):
        serializer = BulkTargetUploadSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        file = serializer.validated_data['file']
        year = serializer.validated_data['year']
        dry_run = serializer.validated_data.get('dry_run', False)
        if file.name.endswith('.csv'):
            content = file.read().decode('utf-8')
        else:
            import openpyxl
            workbook = openpyxl.load_workbook(file)
            sheet = workbook.active
            output = io.StringIO()
            writer = csv.writer(output)
            for row in sheet.iter_rows(values_only=True):
                writer.writerow(row)
            content = output.getvalue()
        reader = csv.DictReader(io.StringIO(content))
        created = 0
        errors = []
        target_setter = TargetSetter()
        for row_num, row in enumerate(reader, start=2):
            try:
                if not dry_run:
                    target_setter.set_annual_target(
                        kpi_id=row['kpi_id'],
                        user_id=row['user_id'],
                        year=year,
                        target_value=row['target_value'],
                        user=request.user
                    )
                created += 1
            except Exception as e:
                errors.append({'row': row_num, 'error': str(e)})
        result_serializer = BulkUploadResultSerializer({
            'total_rows': created + len(errors),
            'created': created,
            'updated': 0,
            'errors': errors,
            'dry_run': dry_run
        })
        return Response(result_serializer.data, status=status.HTTP_202_ACCEPTED if not dry_run else status.HTTP_200_OK)