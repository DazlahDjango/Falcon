# apps/reviews/services/reporting/export_service.py
"""
Export Service - Handles PDF, Excel, CSV exports for reports
"""

import csv
import io
import json
from datetime import datetime
from decimal import Decimal
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from django.http import HttpResponse
from django.utils import timezone
from weasyprint import HTML, CSS
from django.template.loader import render_to_string

from ...constants import ReportFormat
from ..base_service import BaseReviewService


class ExportService(BaseReviewService):
    """
    Handles export of reports to various formats.
    """
    
    @staticmethod
    def export_to_csv(data, filename=None):
        """
        Export data to CSV format.
        
        Args:
            data: List of dictionaries or QuerySet
            filename: Optional filename
        
        Returns:
            HttpResponse: CSV file response
        """
        if filename is None:
            filename = f'report_{timezone.now().strftime("%Y%m%d_%H%M%S")}.csv'
        
        # Convert QuerySet to list of dicts if needed
        if hasattr(data, 'values'):
            data = list(data.values())
        
        if not data:
            return HttpResponse('No data to export', status=404)
        
        # Create CSV response
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        writer = csv.writer(response)
        
        # Write headers
        headers = list(data[0].keys())
        writer.writerow(headers)
        
        # Write data
        for row in data:
            writer.writerow([ExportService._serialize_value(row.get(h)) for h in headers])
        
        return response
    
    @staticmethod
    def export_to_excel(data, sheet_name='Report', filename=None):
        """
        Export data to Excel format.
        
        Args:
            data: List of dictionaries or QuerySet
            sheet_name: Name of Excel sheet
            filename: Optional filename
        
        Returns:
            HttpResponse: Excel file response
        """
        if filename is None:
            filename = f'report_{timezone.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        
        # Convert QuerySet to list of dicts if needed
        if hasattr(data, 'values'):
            data = list(data.values())
        
        if not data:
            return HttpResponse('No data to export', status=404)
        
        wb = Workbook()
        ws = wb.active
        ws.title = sheet_name[:31]  # Excel sheet name max 31 chars
        
        # Define styles
        header_font = Font(bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='4F81BD', end_color='4F81BD', fill_type='solid')
        header_alignment = Alignment(horizontal='center', vertical='center')
        
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Write headers
        headers = list(data[0].keys())
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = border
        
        # Write data
        for row_idx, row in enumerate(data, 2):
            for col_idx, header in enumerate(headers, 1):
                value = ExportService._serialize_value(row.get(header))
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.border = border
        
        # Auto-size columns
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Save to response
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        wb.save(response)
        
        return response
    
    @staticmethod
    def export_to_pdf(template_name, context, filename=None):
        """
        Export HTML template to PDF.
        
        Args:
            template_name: Django template path
            context: Template context dictionary
            filename: Optional filename
        
        Returns:
            HttpResponse: PDF file response
        """
        if filename is None:
            filename = f'report_{timezone.now().strftime("%Y%Y%m%d_%H%M%S")}.pdf'
        
        # Render HTML template
        html_string = render_to_string(template_name, context)
        
        # CSS for PDF styling
        css = CSS(string='''
            @page {
                size: A4;
                margin: 2cm;
            }
            body {
                font-family: Arial, sans-serif;
                font-size: 12px;
                line-height: 1.4;
                color: #333;
            }
            h1 {
                font-size: 24px;
                color: #2c3e50;
                border-bottom: 2px solid #3498db;
                padding-bottom: 10px;
            }
            h2 {
                font-size: 18px;
                color: #2c3e50;
                margin-top: 20px;
            }
            table {
                width: 100%;
                border-collapse: collapse;
                margin: 15px 0;
            }
            th {
                background-color: #3498db;
                color: white;
                padding: 10px;
                text-align: left;
            }
            td {
                border: 1px solid #ddd;
                padding: 8px;
            }
            .footer {
                text-align: center;
                font-size: 10px;
                color: #999;
                margin-top: 30px;
            }
        ''')
        
        # Generate PDF
        html = HTML(string=html_string)
        pdf = html.write_pdf(stylesheets=[css])
        
        # Create response
        response = HttpResponse(pdf, content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        return response
    
    @staticmethod
    def export_to_json(data, filename=None):
        """
        Export data to JSON format.
        
        Args:
            data: Data to export
            filename: Optional filename
        
        Returns:
            HttpResponse: JSON file response
        """
        if filename is None:
            filename = f'report_{timezone.now().strftime("%Y%m%d_%H%M%S")}.json'
        
        # Convert QuerySet if needed
        if hasattr(data, 'values'):
            data = list(data.values())
        
        serialized_data = ExportService._serialize_json(data)
        
        response = HttpResponse(
            json.dumps(serialized_data, indent=2, default=str),
            content_type='application/json'
        )
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        return response
    
    @staticmethod
    def _serialize_value(value):
        """
        Serialize a value for CSV/Excel export.
        
        Args:
            value: Value to serialize
        
        Returns:
            str: Serialized value
        """
        if value is None:
            return ''
        if isinstance(value, (datetime, date)):
            return value.isoformat()
        if isinstance(value, Decimal):
            return float(value)
        if isinstance(value, dict) or isinstance(value, list):
            return json.dumps(value)
        return str(value)
    
    @staticmethod
    def _serialize_json(obj):
        """
        Recursively serialize object for JSON export.
        
        Args:
            obj: Object to serialize
        
        Returns:
            JSON-serializable object
        """
        if obj is None:
            return None
        if isinstance(obj, (datetime, date)):
            return obj.isoformat()
        if isinstance(obj, Decimal):
            return float(obj)
        if isinstance(obj, dict):
            return {k: ExportService._serialize_json(v) for k, v in obj.items()}
        if isinstance(obj, (list, tuple)):
            return [ExportService._serialize_json(item) for item in obj]
        return obj