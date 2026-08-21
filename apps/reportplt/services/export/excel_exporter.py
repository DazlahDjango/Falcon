# apps/reportplt/services/export/excel_exporter.py

import io
import uuid
from typing import Dict, Any, List, Optional
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, LineChart, PieChart, Reference
from openpyxl.chart.series import Series
from django.core.files.storage import default_storage
from django.core.files.base import ContentFile
from apps.reportplt.exceptions import ReportExportError

class ExcelExporter:
    """
    Executive-styled Excel Exporter for Falcon PMS.
    Generates rich, publication-grade Excel workbooks with branded headers,
    KPI stat cards, formatted tables, and charts.
    """

    def __init__(self, config: Optional[Dict] = None):
        self.config = config or {}
        self.wb = Workbook()
        self._setup_palette()

    def _setup_palette(self):
        # Colors
        self.NAVY_HEADER = '1E3A8A'
        self.SLATE_DARK = '0F172A'
        self.BLUE_LIGHT = 'E8F0FE'
        self.GREEN_LIGHT = 'DCFCE7'
        self.GREEN_DARK = '15803D'
        self.YELLOW_LIGHT = 'FEF7E0'
        self.YELLOW_DARK = 'D97706'
        self.PURPLE_LIGHT = 'F3E8FD'
        self.PURPLE_DARK = '7E22CE'
        self.BG_ALT = 'F8FAFC'
        self.BORDER_COLOR = 'CBD5E1'

        # Fonts
        self.font_brand = Font(name='Segoe UI', size=16, bold=True, color='1E3A8A')
        self.font_subtitle = Font(name='Segoe UI', size=9, italic=True, color='64748B')
        self.font_title = Font(name='Segoe UI', size=18, bold=True, color='0F172A')
        self.font_meta = Font(name='Segoe UI', size=9, bold=True, color='475569')
        self.font_meta_val = Font(name='Segoe UI', size=9, color='1E293B')
        self.font_section = Font(name='Segoe UI', size=11, bold=True, color='FFFFFF')
        self.font_card_num = Font(name='Segoe UI', size=22, bold=True, color='0F172A')
        self.font_card_title = Font(name='Segoe UI', size=9, bold=True, color='475569')
        self.font_card_sub = Font(name='Segoe UI', size=8, color='64748B')
        self.font_table_hdr = Font(name='Segoe UI', size=10, bold=True, color='FFFFFF')
        self.font_bold = Font(name='Segoe UI', size=10, bold=True, color='0F172A')
        self.font_regular = Font(name='Segoe UI', size=10, color='334155')

        # Fills
        self.fill_header = PatternFill(start_color=self.NAVY_HEADER, end_color=self.NAVY_HEADER, fill_type='solid')
        self.fill_alt = PatternFill(start_color=self.BG_ALT, end_color=self.BG_ALT, fill_type='solid')
        self.fill_status = PatternFill(start_color=self.GREEN_LIGHT, end_color=self.GREEN_LIGHT, fill_type='solid')
        self.fill_card_blue = PatternFill(start_color='F0F7FF', end_color='F0F7FF', fill_type='solid')
        self.fill_card_green = PatternFill(start_color=self.GREEN_LIGHT, end_color=self.GREEN_LIGHT, fill_type='solid')
        self.fill_card_yellow = PatternFill(start_color=self.YELLOW_LIGHT, end_color=self.YELLOW_LIGHT, fill_type='solid')
        self.fill_card_purple = PatternFill(start_color=self.PURPLE_LIGHT, end_color=self.PURPLE_LIGHT, fill_type='solid')

        # Alignments
        self.align_center = Alignment(horizontal='center', vertical='center')
        self.align_left = Alignment(horizontal='left', vertical='center')
        self.align_right = Alignment(horizontal='right', vertical='center')
        self.align_wrap = Alignment(horizontal='left', vertical='center', wrap_text=True)

        # Borders
        self.border_thin = Border(
            left=Side(style='thin', color=self.BORDER_COLOR),
            right=Side(style='thin', color=self.BORDER_COLOR),
            top=Side(style='thin', color=self.BORDER_COLOR),
            bottom=Side(style='thin', color=self.BORDER_COLOR)
        )
        self.border_card_blue = Border(
            left=Side(style='medium', color='2563EB'),
            right=Side(style='thin', color=self.BORDER_COLOR),
            top=Side(style='thin', color=self.BORDER_COLOR),
            bottom=Side(style='thin', color=self.BORDER_COLOR)
        )
        self.border_card_green = Border(
            left=Side(style='medium', color='16A34A'),
            right=Side(style='thin', color=self.BORDER_COLOR),
            top=Side(style='thin', color=self.BORDER_COLOR),
            bottom=Side(style='thin', color=self.BORDER_COLOR)
        )
        self.border_card_yellow = Border(
            left=Side(style='medium', color='D97706'),
            right=Side(style='thin', color=self.BORDER_COLOR),
            top=Side(style='thin', color=self.BORDER_COLOR),
            bottom=Side(style='thin', color=self.BORDER_COLOR)
        )
        self.border_card_purple = Border(
            left=Side(style='medium', color='9333EA'),
            right=Side(style='thin', color=self.BORDER_COLOR),
            top=Side(style='thin', color=self.BORDER_COLOR),
            bottom=Side(style='thin', color=self.BORDER_COLOR)
        )

    def _format_label(self, raw_key: str) -> str:
        """Converts snake_case or raw_key into clean Title Case."""
        if not raw_key:
            return ""
        words = raw_key.replace('_', ' ').replace('-', ' ').split()
        return " ".join(w.capitalize() for w in words)

    def export(self, data: Dict[str, Any], report_name: str, output_path: Optional[str] = None) -> str:
        try:
            self.wb.remove(self.wb.active)
            self._create_executive_summary_sheet(data, report_name)
            self._create_kpi_sheet(data)
            self._create_chart_sheet(data)
            self._create_data_sheet(data)

            buffer = io.BytesIO()
            self.wb.save(buffer)
            buffer.seek(0)

            if output_path:
                with default_storage.open(output_path, 'wb') as f:
                    f.write(buffer.getvalue())
                return output_path

            file_name = f"reports/{uuid.uuid4()}.xlsx"
            path = default_storage.save(file_name, ContentFile(buffer.getvalue()))
            return path
        except Exception as e:
            raise ReportExportError(f"Excel export failed: {str(e)}")

    def _create_executive_summary_sheet(self, data: Dict, report_name: str):
        ws = self.wb.create_sheet("Executive Summary")
        ws.views.sheetView[0].showGridLines = True

        # ----------------------------------------------------
        # 1. TOP HEADER BANNER
        # ----------------------------------------------------
        # Left: Brand
        ws['A2'] = "FALCON PMS"
        ws['A2'].font = self.font_brand
        ws['A3'] = "Performance. People. Progress."
        ws['A3'].font = self.font_subtitle

        # Center: Report Title
        ws['D2'] = f"{report_name.upper()} REPORT"
        ws['D2'].font = self.font_title
        ws['D2'].alignment = self.align_center
        ws.merge_cells('D2:I2')

        ws['D3'] = "Executive Summary"
        ws['D3'].font = Font(name='Segoe UI', size=11, bold=True, color='475569')
        ws['D3'].alignment = self.align_center
        ws.merge_cells('D3:I3')

        # Right: Metadata Card
        ws['K2'] = "Generated On:"
        ws['K2'].font = self.font_meta
        ws['L2'] = datetime.now().strftime('%d %b %Y %H:%M:%S')
        ws['L2'].font = self.font_meta_val

        ws['K3'] = "Generated By:"
        ws['K3'].font = self.font_meta
        ws['L3'] = "Falcon PMS System"
        ws['L3'].font = self.font_meta_val

        # ----------------------------------------------------
        # 2. STATUS BANNER (Row 5-6)
        # ----------------------------------------------------
        status_text = data.get('status', 'Completed').upper()
        summary_text = data.get('executive_summary', 'The onboarding and operational processes are tracked successfully.')

        ws['A5'] = "STATUS"
        ws['A5'].font = Font(name='Segoe UI', size=8, bold=True, color='16A34A')
        ws['A5'].fill = self.fill_status
        ws['A5'].alignment = self.align_center
        ws['A5'].border = self.border_thin

        ws['A6'] = status_text
        ws['A6'].font = Font(name='Segoe UI', size=12, bold=True, color='15803D')
        ws['A6'].fill = self.fill_status
        ws['A6'].alignment = self.align_center
        ws['A6'].border = self.border_thin

        ws['C5'] = summary_text
        ws['C5'].font = Font(name='Segoe UI', size=10, italic=True, color='334155')
        ws['C5'].alignment = self.align_wrap
        ws.merge_cells('C5:L6')

        # ----------------------------------------------------
        # 3. KPI STAT CARDS (Row 8 - 11)
        # ----------------------------------------------------
        metrics = data.get('metrics', {})
        metric_items = list(metrics.items())

        card_configs = [
            {'col_start': 'A', 'col_end': 'C', 'fill': self.fill_card_blue, 'border': self.border_card_blue, 'color': '1E40AF', 'sub': 'Total metric items'},
            {'col_start': 'D', 'col_end': 'F', 'fill': self.fill_card_green, 'border': self.border_card_green, 'color': '15803D', 'sub': 'Currently active'},
            {'col_start': 'G', 'col_end': 'I', 'fill': self.fill_card_yellow, 'border': self.border_card_yellow, 'color': 'D97706', 'sub': 'Resources over quota'},
            {'col_start': 'J', 'col_end': 'L', 'fill': self.fill_card_purple, 'border': self.border_card_purple, 'color': '7E22CE', 'sub': 'Successful operations'},
        ]

        for idx in range(4):
            cfg = card_configs[idx]
            c_start = cfg['col_start']
            c_end = cfg['col_end']
            
            if idx < len(metric_items):
                key, val = metric_items[idx]
                label = self._format_label(key).upper()
                val_str = f"{val}%" if 'rate' in key or 'percentage' in key else str(val)
            else:
                label = f"METRIC {idx+1}"
                val_str = "0"

            # Top label
            cell_lbl = ws[f'{c_start}8']
            cell_lbl.value = label
            cell_lbl.font = Font(name='Segoe UI', size=9, bold=True, color=cfg['color'])
            cell_lbl.alignment = self.align_center

            # Big Number
            cell_num = ws[f'{c_start}9']
            cell_num.value = val_str
            cell_num.font = Font(name='Segoe UI', size=20, bold=True, color='0F172A')
            cell_num.alignment = self.align_center

            # Subtitle
            cell_sub = ws[f'{c_start}11']
            cell_sub.value = cfg['sub']
            cell_sub.font = self.font_card_sub
            cell_sub.alignment = self.align_center

            # Apply fill & merge
            for r in range(8, 12):
                for col in [c_start, chr(ord(c_start)+1), c_end]:
                    ws[f'{col}{r}'].fill = cfg['fill']
                    ws[f'{col}{r}'].border = self.border_thin

        # ----------------------------------------------------
        # 4. TABLES SECTION (Row 13+)
        # ----------------------------------------------------
        # Table 1: Key Metrics (Left Side - Col A:C)
        ws['A13'] = "KEY METRICS"
        ws['A13'].font = self.font_section
        ws['A13'].fill = self.fill_header
        ws['A13'].alignment = self.align_left
        ws.merge_cells('A13:C13')

        ws['A14'] = "Metric Name"
        ws['A14'].font = self.font_table_hdr
        ws['A14'].fill = PatternFill(start_color='334155', end_color='334155', fill_type='solid')
        ws.merge_cells('A14:B14')

        ws['C14'] = "Value"
        ws['C14'].font = self.font_table_hdr
        ws['C14'].fill = PatternFill(start_color='334155', end_color='334155', fill_type='solid')
        ws['C14'].alignment = self.align_right

        curr_row = 15
        for key, val in metrics.items():
            ws[f'A{curr_row}'] = self._format_label(key)
            ws[f'A{curr_row}'].font = self.font_regular
            ws.merge_cells(f'A{curr_row}:B{curr_row}')

            val_str = f"{val}%" if 'rate' in key or 'percentage' in key else str(val)
            ws[f'C{curr_row}'] = val_str
            ws[f'C{curr_row}'].font = self.font_bold
            ws[f'C{curr_row}'].alignment = self.align_right

            if curr_row % 2 == 1:
                ws[f'A{curr_row}'].fill = self.fill_alt
                ws[f'B{curr_row}'].fill = self.fill_alt
                ws[f'C{curr_row}'].fill = self.fill_alt

            ws[f'A{curr_row}'].border = self.border_thin
            ws[f'B{curr_row}'].border = self.border_thin
            ws[f'C{curr_row}'].border = self.border_thin
            curr_row += 1

        # Table 2: Report Information (Col A:C below Key Metrics)
        curr_row += 1
        ws[f'A{curr_row}'] = "REPORT INFORMATION"
        ws[f'A{curr_row}'].font = self.font_section
        ws[f'A{curr_row}'].fill = self.fill_header
        ws.merge_cells(f'A{curr_row}:C{curr_row}')
        curr_row += 1

        info_items = [
            ("Report Name", report_name),
            ("Generation Time", datetime.now().strftime('%d %b %Y %H:%M:%S')),
            ("Status", status_text),
            ("Report Scope", "All Active Tenants / System Scope"),
            ("Data As Of", datetime.now().strftime('%d %b %Y')),
        ]

        for label, val in info_items:
            ws[f'A{curr_row}'] = label
            ws[f'A{curr_row}'].font = self.font_regular

            ws[f'B{curr_row}'] = val
            ws[f'B{curr_row}'].font = self.font_bold
            ws.merge_cells(f'B{curr_row}:C{curr_row}')

            if curr_row % 2 == 1:
                ws[f'A{curr_row}'].fill = self.fill_alt
                ws[f'B{curr_row}'].fill = self.fill_alt
                ws[f'C{curr_row}'].fill = self.fill_alt

            ws[f'A{curr_row}'].border = self.border_thin
            ws[f'B{curr_row}'].border = self.border_thin
            ws[f'C{curr_row}'].border = self.border_thin
            curr_row += 1

        # Table 3: KPI Summary Table (Right Side - Col E:L)
        ws['E13'] = "KPI SUMMARY & BREAKDOWN"
        ws['E13'].font = self.font_section
        ws['E13'].fill = self.fill_header
        ws.merge_cells('E13:L13')

        kpi_hdrs = [("KPI Status", 'E14:F14'), ("Count", 'G14'), ("Percentage", 'H14'), ("Visual Progress", 'I14:L14')]
        for title, cell_range in kpi_hdrs:
            if ':' in cell_range:
                start_c = cell_range.split(':')[0]
                ws[start_c] = title
                ws[start_c].font = self.font_table_hdr
                ws[start_c].fill = PatternFill(start_color='334155', end_color='334155', fill_type='solid')
                ws.merge_cells(cell_range)
            else:
                ws[cell_range] = title
                ws[cell_range].font = self.font_table_hdr
                ws[cell_range].fill = PatternFill(start_color='334155', end_color='334155', fill_type='solid')

        kpi_rows = [
            ("Total KPIs", len(data.get('kpis', [])), "100%", "████████████████████ 100%"),
            ("On Track", sum(1 for k in data.get('kpis', []) if k.get('status') == 'On Track'), "0%", "░░░░░░░░░░░░░░░░░░░░ 0%"),
            ("At Risk", sum(1 for k in data.get('kpis', []) if k.get('status') == 'At Risk'), "0%", "░░░░░░░░░░░░░░░░░░░░ 0%"),
            ("Off Track", sum(1 for k in data.get('kpis', []) if k.get('status') == 'Off Track'), "0%", "░░░░░░░░░░░░░░░░░░░░ 0%"),
        ]

        k_row = 15
        for st_lbl, cnt, pct, prog in kpi_rows:
            ws[f'E{k_row}'] = st_lbl
            ws[f'E{k_row}'].font = self.font_regular
            ws.merge_cells(f'E{k_row}:F{k_row}')

            ws[f'G{k_row}'] = cnt
            ws[f'G{k_row}'].font = self.font_bold
            ws[f'G{k_row}'].alignment = self.align_center

            ws[f'H{k_row}'] = pct
            ws[f'H{k_row}'].font = self.font_regular
            ws[f'H{k_row}'].alignment = self.align_center

            ws[f'I{k_row}'] = prog
            ws[f'I{k_row}'].font = Font(name='Segoe UI', size=9, color='2563EB')
            ws.merge_cells(f'I{k_row}:L{k_row}')

            for c in ['E','F','G','H','I','J','K','L']:
                ws[f'{c}{k_row}'].border = self.border_thin

            k_row += 1

        # ----------------------------------------------------
        # 5. FOOTER BANNER
        # ----------------------------------------------------
        footer_row = max(curr_row, k_row) + 2
        ws[f'A{footer_row}'] = "Confidential Report | FALCON PMS Powered by FalconTech Ltd. All rights reserved."
        ws[f'A{footer_row}'].font = Font(name='Segoe UI', size=9, italic=True, color='94A3B8')
        ws[f'A{footer_row}'].alignment = self.align_center
        ws.merge_cells(f'A{footer_row}:L{footer_row}')

        # Column widths
        ws.column_dimensions['A'].width = 28
        ws.column_dimensions['B'].width = 24
        ws.column_dimensions['C'].width = 16
        ws.column_dimensions['D'].width = 16
        ws.column_dimensions['E'].width = 18
        ws.column_dimensions['F'].width = 18
        ws.column_dimensions['G'].width = 14
        ws.column_dimensions['H'].width = 14
        ws.column_dimensions['I'].width = 16
        ws.column_dimensions['J'].width = 16
        ws.column_dimensions['K'].width = 18
        ws.column_dimensions['L'].width = 24

    def _create_kpi_sheet(self, data: Dict):
        ws = self.wb.create_sheet("KPI Details")
        ws.views.sheetView[0].showGridLines = True

        headers = ['KPI Name', 'Target', 'Actual', 'Progress', 'Status']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = self.font_table_hdr
            cell.fill = self.fill_header
            cell.alignment = self.align_center
            cell.border = self.border_thin

        kpis = data.get('kpis', [])
        if not kpis:
            ws.cell(row=2, column=1, value="No detailed KPI records available for this report run.").font = self.font_regular
            ws.merge_cells('A2:E2')
            return

        for row_idx, kpi in enumerate(kpis, 2):
            ws.cell(row=row_idx, column=1, value=kpi.get('name', '')).alignment = self.align_left
            ws.cell(row=row_idx, column=2, value=kpi.get('target', 0)).alignment = self.align_right
            ws.cell(row=row_idx, column=3, value=kpi.get('actual', 0)).alignment = self.align_right
            ws.cell(row=row_idx, column=4, value=f"{kpi.get('progress', 0)}%").alignment = self.align_right

            status_cell = ws.cell(row=row_idx, column=5, value=kpi.get('status', 'Completed'))
            status_cell.alignment = self.align_center
            status_cell.font = Font(name='Segoe UI', size=10, bold=True)
            status_cell.fill = self.fill_status
            
            for c in range(1, 6):
                ws.cell(row=row_idx, column=c).border = self.border_thin

        for col in range(1, 6):
            ws.column_dimensions[get_column_letter(col)].width = 22

    def _create_chart_sheet(self, data: Dict):
        ws = self.wb.create_sheet("Visual Charts")
        ws.views.sheetView[0].showGridLines = True

        charts = data.get('charts', [])
        if not charts:
            ws['A1'] = "Visual charts are generated automatically upon report compilation."
            ws['A1'].font = self.font_regular
            return

        chart_data = charts[0].get('data', {})
        labels = chart_data.get('labels', [])
        values = chart_data.get('values', [])
        if not labels or not values:
            ws['A1'] = "No chart dataset available"
            return

        ws.cell(row=1, column=1, value="Category").font = self.font_table_hdr
        ws.cell(row=1, column=2, value="Value").font = self.font_table_hdr

        for idx, (label, value) in enumerate(zip(labels, values), 2):
            ws.cell(row=idx, column=1, value=label)
            ws.cell(row=idx, column=2, value=value)

        chart = BarChart()
        chart.title = charts[0].get('title', 'Performance Metrics Summary')
        chart.x_axis.title = "Categories"
        chart.y_axis.title = "Count"
        
        data_ref = Reference(ws, min_col=2, min_row=1, max_row=len(values)+1)
        labels_ref = Reference(ws, min_col=1, min_row=2, max_row=len(values)+1)
        chart.add_data(data_ref, titles_from_data=True)
        chart.set_categories(labels_ref)
        chart.height = 12
        chart.width = 18
        ws.add_chart(chart, "D2")

    def _create_data_sheet(self, data: Dict):
        ws = self.wb.create_sheet("Raw Data")
        ws.views.sheetView[0].showGridLines = True

        tables = data.get('tables', [])
        if not tables:
            ws['A1'] = "No detailed raw data rows available for this report."
            ws['A1'].font = self.font_regular
            return

        table_data = tables[0]
        columns = table_data.get('columns', [])
        rows = table_data.get('rows', [])
        if not columns or not rows:
            ws['A1'] = "No detailed raw data rows available for this report."
            return

        for col, header in enumerate(columns, 1):
            cell = ws.cell(row=1, column=col, value=self._format_label(header))
            cell.font = self.font_table_hdr
            cell.fill = self.fill_header
            cell.alignment = self.align_center
            cell.border = self.border_thin

        for row_idx, row in enumerate(rows[:1000], 2):
            for col_idx, value in enumerate(row, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.border = self.border_thin
                cell.font = self.font_regular

        for col in range(1, len(columns) + 1):
            ws.column_dimensions[get_column_letter(col)].width = 20

    def export_to_bytes(self, data: Dict[str, Any], report_name: str = "Report", config: Optional[Dict] = None) -> bytes:
        if config:
            self.config.update(config)
        self.export(data, report_name)
        buffer = io.BytesIO()
        self.wb.save(buffer)
        buffer.seek(0)
        return buffer.getvalue()